from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait 
from selenium.webdriver.support import expected_conditions as ec 
import pytest
import openpyxl
from constants import globalConstants as c


class Test_Tobeto_Register():
  def setup_method(self):
    self.driver = webdriver.Chrome()
    self.driver.get(c.BASE_URL)
    self.driver.maximize_window()
    
  
  def teardown_method(self):
    self.driver.quit()


  def getData():
      excel = openpyxl.load_workbook(c.Register_xlsx)
      sheet = excel["Sheet1"] 
      rows = sheet.max_row 
      data = []
      for i in range(2,rows+1):
          firstName = sheet.cell(i,1).value
          lastName = sheet.cell(i,2).value
          email = sheet.cell(i,3).value
          password = sheet.cell(i,4).value
          passwordAgain = sheet.cell(i,5).value
          phoneNumber = sheet.cell(i,6).value
          data.append((firstName,lastName,email,password,passwordAgain,phoneNumber))

      return data
  
  def getData1():
      excel = openpyxl.load_workbook(c.Register1_xlsx)
      sheet = excel["Sheet1"] 
      rows = sheet.max_row 
      data = []
      for i in range(2,rows+1):
          firstName = sheet.cell(i,1).value
          lastName = sheet.cell(i,2).value
          email = sheet.cell(i,3).value
          password = sheet.cell(i,4).value
          passwordAgain = sheet.cell(i,5).value
          data.append((firstName,lastName,email,password,passwordAgain))

      return data
  
  def getData2():
      excel = openpyxl.load_workbook(c.Register2_xlsx)
      sheet = excel["Sheet1"] 
      rows = sheet.max_row 
      data = []
      for i in range(2,rows+1):
          firstName = sheet.cell(i,1).value
          lastName = sheet.cell(i,2).value
          email = sheet.cell(i,3).value
          phoneNumber= sheet.cell(i,4).value
          
          data.append((firstName,lastName,email,phoneNumber))

      return data


  def test_invalidEmail(self):
    registerId = self.driver.find_element(By.LINK_TEXT, c.REGISTER)
    registerId .click()
    emailName = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, "email")))
    emailName.click
    emailName.send_keys("abcd")
    errorMessage = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, c.INVALID_EMAIL_XPATH))) 
    assert errorMessage.text == c.INVALID_EMAIL 
  

  @pytest.mark.parametrize("firstName,lastName,email,password,passwordAgain,phoneNumber", getData()) 
  def test_invalidRegister(self,firstName,lastName,email,password,passwordAgain,phoneNumber):
    registerId = self.driver.find_element(By.LINK_TEXT, c.REGISTER)
    registerId .click()
    firstnameId = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.FIRSTNAME_NAME)))
    firstnameId.click()
    firstnameId.send_keys(firstName)
    lastnameId = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.LASTNAME_NAME)))
    lastnameId.click()
    lastnameId.send_keys(lastName)
    emailName = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.EMAIL_NAME)))
    emailName.click
    emailName.send_keys(email)
    self.driver.execute_script("window.scrollTo(0,600)")
    enterPassword = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.PASSWORD_NAME)))
    enterPassword.click()
    enterPassword.send_keys(password)
    self.driver.execute_script("window.scrollTo(0,600)")
    enterPasswordAgain = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.PASSWORDAGAIN_NAME)))
    enterPasswordAgain.click()
    enterPasswordAgain.send_keys(passwordAgain)
    self.driver.execute_script("window.scrollTo(0,600)")
    sleep(5)
    registerButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.REGISTER_XPATH)))
    registerButton.click()
    self.driver.find_element(By.NAME, "contact").click()
    self.driver.find_element(By.NAME, "membershipContrat").click()
    self.driver.find_element(By.NAME, "emailConfirmation").click()
    self.driver.find_element(By.NAME, "phoneConfirmation").click()
    enterPhoneNumber = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.ID, c.PHONE_NUMBER_ID)))
    enterPhoneNumber.click()
    enterPhoneNumber.send_keys(phoneNumber)
    WebDriverWait(self.driver, 10).until(ec.frame_to_be_available_and_switch_to_it((By.XPATH, "//iframe[starts-with(@name, 'a-') and starts-with(@src, 'https://www.google.com/recaptcha')]")))
    recaptcha = WebDriverWait(self.driver, 20).until(ec.element_to_be_clickable((By.CLASS_NAME, "recaptcha-checkbox-border")))
    recaptcha.click()
    self.driver.switch_to.default_content()
    sleep(60)
    finalRegister = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, c.FINAL_REGISTER_XPATH)))
    finalRegister.click()
    sleep(3)
    cantRegister = self.driver.find_element(By.CLASS_NAME , "toast-body")
    assert cantRegister.text == c.CANT_REGISTER
    self.driver.close()


  @pytest.mark.parametrize("firstName,lastName,email,password,passwordAgain", getData1()) 
  def test_minPhoneNumber(self,firstName,lastName,email,password,passwordAgain):
    registerId = self.driver.find_element(By.LINK_TEXT, c.REGISTER)
    registerId .click()
    firstnameId = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.FIRSTNAME_NAME)))
    firstnameId.click()
    firstnameId.send_keys(firstName)
    lastnameId = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.LASTNAME_NAME)))
    lastnameId.click()
    lastnameId.send_keys(lastName)
    emailName = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.EMAIL_NAME)))
    emailName.click
    emailName.send_keys(email)
    self.driver.execute_script("window.scrollTo(0,500)")
    enterPassword = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.PASSWORD_NAME)))
    enterPassword.click()
    enterPassword.send_keys(password)
    self.driver.execute_script("window.scrollTo(0,500)")
    enterPasswordAgain = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.PASSWORDAGAIN_NAME)))
    enterPasswordAgain.click()
    enterPasswordAgain.send_keys(passwordAgain)
    self.driver.execute_script("window.scrollTo(0,500)")
    sleep(3)
    registerButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.REGISTER_XPATH)))
    registerButton.click()
    self.driver.find_element(By.NAME, "contact").click()
    self.driver.find_element(By.NAME, "membershipContrat").click()
    self.driver.find_element(By.NAME, "emailConfirmation").click()
    self.driver.find_element(By.NAME, "phoneConfirmation").click()
    enterPhoneNumber = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.ID, c.PHONE_NUMBER_ID)))
    enterPhoneNumber.click()
    enterPhoneNumber.send_keys("12345")
    WebDriverWait(self.driver, 10).until(ec.frame_to_be_available_and_switch_to_it((By.XPATH, "//iframe[starts-with(@name, 'a-') and starts-with(@src, 'https://www.google.com/recaptcha')]")))
    recaptcha = WebDriverWait(self.driver, 20).until(ec.element_to_be_clickable((By.CLASS_NAME, "recaptcha-checkbox-border")))
    recaptcha.click()
    self.driver.switch_to.default_content()
    sleep(30)
    finalRegister = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, c.FINAL_REGISTER_XPATH)))
    finalRegister.click()
    minPhoneNumber = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, c.MIN_PHONE_NUMBER_XPATH)))
    #minPhoneNumber = self.driver.find_element((By.XPATH, c.MIN_PHONE_NUMBER_XPATH))
    assert minPhoneNumber.text == c.MIN_PHONE_NUMBER
    self.driver.close()

  @pytest.mark.parametrize("firstName,lastName,email,password,passwordAgain", getData1()) 
  def test_maxPhoneNumber(self,firstName,lastName,email,password,passwordAgain):
    registerId = self.driver.find_element(By.LINK_TEXT, c.REGISTER)
    registerId .click()
    firstnameId = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.FIRSTNAME_NAME)))
    firstnameId.click()
    firstnameId.send_keys(firstName)
    lastnameId = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.LASTNAME_NAME)))
    lastnameId.click()
    lastnameId.send_keys(lastName)
    emailName = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.EMAIL_NAME)))
    emailName.click
    emailName.send_keys(email)
    self.driver.execute_script("window.scrollTo(0,500)")
    enterPassword = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.PASSWORD_NAME)))
    enterPassword.click()
    enterPassword.send_keys(password)
    self.driver.execute_script("window.scrollTo(0,500)")
    enterPasswordAgain = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.PASSWORDAGAIN_NAME)))
    enterPasswordAgain.click()
    enterPasswordAgain.send_keys(passwordAgain)
    self.driver.execute_script("window.scrollTo(0,500)")
    registerButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.REGISTER_XPATH)))
    registerButton.click()
    self.driver.find_element(By.NAME, "contact").click()
    self.driver.find_element(By.NAME, "membershipContrat").click()
    self.driver.find_element(By.NAME, "emailConfirmation").click()
    self.driver.find_element(By.NAME, "phoneConfirmation").click()
    enterPhoneNumber = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.ID, c.PHONE_NUMBER_ID)))
    enterPhoneNumber.click()
    enterPhoneNumber.send_keys("12345678911")
    WebDriverWait(self.driver, 10).until(ec.frame_to_be_available_and_switch_to_it((By.XPATH, "//iframe[starts-with(@name, 'a-') and starts-with(@src, 'https://www.google.com/recaptcha')]")))
    recaptcha = WebDriverWait(self.driver, 20).until(ec.element_to_be_clickable((By.CLASS_NAME, "recaptcha-checkbox-border")))
    recaptcha.click()
    self.driver.switch_to.default_content()
    sleep(30)
    finalRegister = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, c.FINAL_REGISTER_XPATH)))
    finalRegister.click()
    maxPhoneNumber = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, c.MAX_PHONE_NUMBER_XPATH)))
    assert maxPhoneNumber.text == c.MAX_PHONE_NUMBER
    self.driver.close()
  

  @pytest.mark.parametrize("firstName,lastName,email,password,passwordAgain", getData1()) 
  def test_emptyPhoneNumber(self,firstName,lastName,email,password,passwordAgain):
    registerId = self.driver.find_element(By.LINK_TEXT, c.REGISTER)
    registerId .click()
    firstnameId = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.FIRSTNAME_NAME)))
    firstnameId.click()
    firstnameId.send_keys(firstName)
    lastnameId = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.LASTNAME_NAME)))
    lastnameId.click()
    lastnameId.send_keys(lastName)
    emailName = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.EMAIL_NAME)))
    emailName.click
    emailName.send_keys(email)
    self.driver.execute_script("window.scrollTo(0,500)")
    enterPassword = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.PASSWORD_NAME)))
    enterPassword.click()
    enterPassword.send_keys(password)
    self.driver.execute_script("window.scrollTo(0,500)")
    enterPasswordAgain = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.PASSWORDAGAIN_NAME)))
    enterPasswordAgain.click()
    enterPasswordAgain.send_keys(passwordAgain)
    self.driver.execute_script("window.scrollTo(0,500)")
    registerButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.REGISTER_XPATH)))
    registerButton.click()
    self.driver.find_element(By.NAME, "contact").click()
    self.driver.find_element(By.NAME, "membershipContrat").click()
    self.driver.find_element(By.NAME, "emailConfirmation").click()
    self.driver.find_element(By.NAME, "phoneConfirmation").click()
    enterPhoneNumber = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.ID, c.PHONE_NUMBER_ID)))
    enterPhoneNumber.click()
    enterPhoneNumber.send_keys(" ")
    WebDriverWait(self.driver, 10).until(ec.frame_to_be_available_and_switch_to_it((By.XPATH, "//iframe[starts-with(@name, 'a-') and starts-with(@src, 'https://www.google.com/recaptcha')]")))
    recaptcha = WebDriverWait(self.driver, 20).until(ec.element_to_be_clickable((By.CLASS_NAME, "recaptcha-checkbox-border")))
    recaptcha.click()
    self.driver.switch_to.default_content()
    sleep(30)
    finalRegister = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, c.FINAL_REGISTER_XPATH)))
    finalRegister.click()
    sleep(10)
    emptyPhoneNumber = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, c.EMPTY_PHONE_NUMBER_XPATH)))
    assert emptyPhoneNumber.text == c.EMPTY_PHONE_NUMBER
    self.driver.close()

  @pytest.mark.parametrize("firstName,lastName,email,phoneNumber", getData2()) 
  def test_Missing_Password(self,firstName,lastName,email,phoneNumber):
    registerId = self.driver.find_element(By.LINK_TEXT, c.REGISTER)
    registerId .click()
    firstnameId = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.FIRSTNAME_NAME)))
    firstnameId.click()
    firstnameId.send_keys(firstName)
    lastnameId = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.LASTNAME_NAME)))
    lastnameId.click()
    lastnameId.send_keys(lastName)
    emailName = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.EMAIL_NAME)))
    emailName.click
    emailName.send_keys(email)
    self.driver.execute_script("window.scrollTo(0,600)")
    enterPassword = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME,"password")))
    enterPassword.click()
    enterPassword.send_keys("12345")
    self.driver.execute_script("window.scrollTo(0,600)")
    enterPasswordAgain = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, "passwordAgain")))
    enterPasswordAgain.click()
    enterPasswordAgain.send_keys("12345")
    self.driver.execute_script("window.scrollTo(0,600)")
    sleep(5)
    registerButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.REGISTER_XPATH)))
    registerButton.click()
    self.driver.find_element(By.NAME, "contact").click()
    self.driver.find_element(By.NAME, "membershipContrat").click()
    self.driver.find_element(By.NAME, "emailConfirmation").click()
    self.driver.find_element(By.NAME, "phoneConfirmation").click()
    enterPhoneNumber = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.ID, c.PHONE_NUMBER_ID)))
    enterPhoneNumber.click()
    enterPhoneNumber.send_keys(phoneNumber)
    WebDriverWait(self.driver, 10).until(ec.frame_to_be_available_and_switch_to_it((By.XPATH, "//iframe[starts-with(@name, 'a-') and starts-with(@src, 'https://www.google.com/recaptcha')]")))
    recaptcha = WebDriverWait(self.driver, 20).until(ec.element_to_be_clickable((By.CLASS_NAME, "recaptcha-checkbox-border")))
    recaptcha.click()
    self.driver.switch_to.default_content()
    sleep(30)
    finalRegister = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, c.FINAL_REGISTER_XPATH)))
    finalRegister.click()
    sleep(5)
    missing_password = self.driver.find_element(By.CLASS_NAME , "toast-body")
    assert missing_password.text == c.MISSING_PASSWORD
    sleep(5)
    self.driver.close()




  @pytest.mark.parametrize("firstName,lastName,email,phoneNumber", getData2()) 
  def test_Unmatched_Password(self,firstName,lastName,email,phoneNumber):
    registerId = self.driver.find_element(By.LINK_TEXT, c.REGISTER)
    registerId .click()
    firstnameId = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.FIRSTNAME_NAME)))
    firstnameId.click()
    firstnameId.send_keys(firstName)
    lastnameId = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.LASTNAME_NAME)))
    lastnameId.click()
    lastnameId.send_keys(lastName)
    emailName = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.EMAIL_NAME)))
    emailName.click
    emailName.send_keys(email)
    self.driver.execute_script("window.scrollTo(0,600)")
    enterPassword = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME,"password")))
    enterPassword.click()
    enterPassword.send_keys("12345678")
    self.driver.execute_script("window.scrollTo(0,600)")
    enterPasswordAgain = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, "passwordAgain")))
    enterPasswordAgain.click()
    enterPasswordAgain.send_keys("123456")
    self.driver.execute_script("window.scrollTo(0,600)")
    sleep(5)
    registerButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.REGISTER_XPATH)))
    registerButton.click()
    self.driver.find_element(By.NAME, "contact").click()
    self.driver.find_element(By.NAME, "membershipContrat").click()
    self.driver.find_element(By.NAME, "emailConfirmation").click()
    self.driver.find_element(By.NAME, "phoneConfirmation").click()
    enterPhoneNumber = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.ID, c.PHONE_NUMBER_ID)))
    enterPhoneNumber.click()
    enterPhoneNumber.send_keys(phoneNumber)
    WebDriverWait(self.driver, 10).until(ec.frame_to_be_available_and_switch_to_it((By.XPATH, "//iframe[starts-with(@name, 'a-') and starts-with(@src, 'https://www.google.com/recaptcha')]")))
    recaptcha = WebDriverWait(self.driver, 20).until(ec.element_to_be_clickable((By.CLASS_NAME, "recaptcha-checkbox-border")))
    recaptcha.click()
    self.driver.switch_to.default_content()
    sleep(60)
    finalRegister = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, c.FINAL_REGISTER_XPATH)))
    finalRegister.click()
    sleep(5)
    unmatched_password = self.driver.find_element(By.CLASS_NAME , "toast-body")
    assert unmatched_password.text == c.UNMATCHED_PASSWORD
    self.driver.close()


  @pytest.mark.parametrize("firstName,lastName,email,phoneNumber", getData2()) 
  def test_Invalid_Informations(self,firstName,lastName,email,phoneNumber):
    register_id = self.driver.find_element(By.LINK_TEXT, c.REGISTER)
    register_id .click()
    firstname_id = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.FIRSTNAME_NAME)))
    firstname_id.click()
    firstname_id.send_keys(firstName)
    lastname_id = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.LASTNAME_NAME)))
    lastname_id.click()
    lastname_id.send_keys(lastName)
    email_name = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.EMAIL_NAME)))
    email_name.click
    email_name.send_keys(email)
    self.driver.execute_script("window.scrollTo(0,600)")
    enter_password = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME,"password")))
    enter_password.click()
    enter_password.send_keys("12345678")
    self.driver.execute_script("window.scrollTo(0,600)")
    enter_password_again = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, "passwordAgain")))
    enter_password_again.click()
    enter_password_again.send_keys("123456")
    self.driver.execute_script("window.scrollTo(0,600)")
    sleep(5)
    register_button = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.REGISTER_XPATH)))
    register_button.click()
    self.driver.find_element(By.NAME, "contact").click()
    self.driver.find_element(By.NAME, "membershipContrat").click()
    self.driver.find_element(By.NAME, "emailConfirmation").click()
    self.driver.find_element(By.NAME, "phoneConfirmation").click()
    enter_phone_number = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.ID, c.PHONE_NUMBER_ID)))
    enter_phone_number.click()
    enter_phone_number.send_keys(phoneNumber)
    WebDriverWait(self.driver, 10).until(ec.frame_to_be_available_and_switch_to_it((By.XPATH, "//iframe[starts-with(@name, 'a-') and starts-with(@src, 'https://www.google.com/recaptcha')]")))
    recaptcha = WebDriverWait(self.driver, 20).until(ec.element_to_be_clickable((By.CLASS_NAME, "recaptcha-checkbox-border")))
    recaptcha.click()
    self.driver.switch_to.default_content()
    sleep(60)
    final_register = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, c.FINAL_REGISTER_XPATH)))
    final_register.click()
    sleep(5)
    invalid_informations = self.driver.find_element(By.CLASS_NAME , "toast-body")
    assert invalid_informations.text == c.INVALID_INFORMATIONS
    self.driver.close()