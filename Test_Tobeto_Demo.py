from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait 
from selenium.webdriver.support import expected_conditions as ec 
import pytest
import openpyxl
from constants1 import globalConstants1 as c


class Test_Tobeto_Demo():
  def setup_method(self):
    self.driver = webdriver.Chrome()
    self.driver.get(c.BASE_URL)
    self.driver.maximize_window()
    
  
  def teardown_method(self):
    self.driver.quit()


  def getData():
      excel = openpyxl.load_workbook(c.Login_xlsx)
      sheet = excel["Sheet1"] 
      rows = sheet.max_row 
      data = []
      for i in range(2,rows+1):
          email = sheet.cell(i,1).value
          password = sheet.cell(i,2).value
          data.append((email,password))
      
      return data

  def getData1():
      excel = openpyxl.load_workbook(c.Register_xlsx)
      sheet = excel["Sheet1"] 
      rows = sheet.max_row 
      data = []
      for i in range(2,rows+1):
          firstName = sheet.cell(i,1).value
          lastName = sheet.cell(i,2).value
          email = sheet.cell(i,3).value
          password = sheet.cell(i,4).value
          passwordAgain = sheet.cell(i,5).value
          phoneNumber = sheet.cell(i,6).value
          data.append((firstName,lastName,email,password,passwordAgain,phoneNumber))

      return data


  @pytest.mark.parametrize("email,password",[("","")])
  def test_emptylogin(self,email,password):
    emailInput = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.EMAIL_NAME)))
    emailInput.click()
    emailInput.send_keys(email)
    passwordInput = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.PASSWORD_NAME)))
    passwordInput.click()
    passwordInput.send_keys(password)
    loginButton = self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_XPATH)
    loginButton.click()
    errorMessage1 = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.EMPTYMESSAGE_XPATH_1))) 
    assert errorMessage1.text == c.EMPTYMESSAGE
    errorMessage2 = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.EMPTYMESSAGE_XPATH_2))) 
    assert errorMessage2.text == c.EMPTYMESSAGE
    self.driver.close()


  def test_invalidEmail(self):
    registerId = self.driver.find_element(By.LINK_TEXT, c.REGISTER)
    registerId .click()
    emailName = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, "email")))
    emailName.click
    emailName.send_keys("abcd")
    errorMessage = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, c.INVALID_EMAIL_XPATH))) 
    assert errorMessage.text == c.INVALID_EMAIL 
    


  @pytest.mark.parametrize("email,password",getData())
  def test_valid_Login(self,email,password):
    emailInput = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.EMAIL_NAME)))
    emailInput.click()
    emailInput.send_keys(email)
    passwordInput = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.PASSWORD_NAME)))
    passwordInput.click()
    passwordInput.send_keys(password)
    loginButton = self.driver.find_element(By.XPATH, c.LOGIN_BUTTON_XPATH)
    loginButton.click()
    validLogin = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, c.VALID_LOGIN_XPATH)))
    assert validLogin.text == c.VALID_LOGIN
    self.driver.close()
  

  @pytest.mark.parametrize("firstName,lastName,email,password,passwordAgain,phoneNumber", getData1()) 
  def test_invalidRegister(self,firstName,lastName,email,password,passwordAgain,phoneNumber):
    registerId = self.driver.find_element(By.LINK_TEXT, c.REGISTER)
    registerId .click()
    firstnameId = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.FIRSTNAME_NAME)))
    firstnameId.click()
    firstnameId.send_keys(firstName)
    lastnameId = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.LASTNAME_NAME)))
    lastnameId.click()
    lastnameId.send_keys(lastName)
    emailName = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.EMAIL_NAME)))
    emailName.click
    emailName.send_keys(email)
    self.driver.execute_script("window.scrollTo(0,500)")
    enterPassword = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.PASSWORD_NAME)))
    enterPassword.click()
    enterPassword.send_keys(password)
    self.driver.execute_script("window.scrollTo(0,500)")
    enterPasswordAgain = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.PASSWORDAGAIN_NAME)))
    enterPasswordAgain.click()
    enterPasswordAgain.send_keys(passwordAgain)
    self.driver.execute_script("window.scrollTo(0,500)")
    registerButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.REGISTER_XPATH)))
    registerButton.click()
    self.driver.find_element(By.NAME, "contact").click()
    self.driver.find_element(By.NAME, "membershipContrat").click()
    self.driver.find_element(By.NAME, "emailConfirmation").click()
    self.driver.find_element(By.NAME, "phoneConfirmation").click()
    enterPhoneNumber = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.ID, c.PHONE_NUMBER_ID)))
    enterPhoneNumber.click()
    enterPhoneNumber.send_keys(phoneNumber)
    WebDriverWait(self.driver, 10).until(ec.frame_to_be_available_and_switch_to_it((By.XPATH, "//iframe[starts-with(@name, 'a-') and starts-with(@src, 'https://www.google.com/recaptcha')]")))
    recaptcha = WebDriverWait(self.driver, 20).until(ec.element_to_be_clickable((By.CLASS_NAME, "recaptcha-checkbox-border")))
    recaptcha.click()
    #self.driver.switch_to.default_content()
    #self.driver.execute_script("window.scrollTo(0,500)")
    finalRegister = self.driver.find_element(By.XPATH, c.FINAL_REGISTER_XPATH)
    finalRegister.click()
    cantRegister = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, c.CANT_REGISTER_XPATH)))
    assert cantRegister.text == c.CANT_REGISTER
    self.driver.close()



