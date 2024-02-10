from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait 
from selenium.webdriver.support import expected_conditions as ec 
import pytest
import openpyxl
from constants import globalConstants as c

class Test_Tobeto_Login():
  def setup_method(self):
    self.driver = webdriver.Chrome()
    self.driver.get(c.BASE_URL)
    self.driver.maximize_window()
    
  
  def teardown_method(self):
    self.driver.quit()


  def getData():
      excel = openpyxl.load_workbook(c.Login_xlsx)
      sheet = excel["Sheet1"] 
      rows = sheet.max_row 
      data = []
      for i in range(2,rows+1):
          email = sheet.cell(i,1).value
          password = sheet.cell(i,2).value
          data.append((email,password))
      
      return data
  
  def getData1():
      excel = openpyxl.load_workbook(c.Login1_xlsx)
      sheet = excel["Sheet1"] 
      rows = sheet.max_row 
      data = []
      for i in range(2,rows+1):
          email = sheet.cell(i,1).value
          password = sheet.cell(i,2).value
          data.append((email,password))
      
      return data

  
  @pytest.mark.parametrize("email,password",[("","")])
  def test_emptylogin(self,email,password):
    emailInput = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.EMAIL_NAME)))
    emailInput.click()
    emailInput.send_keys(email)
    passwordInput = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.PASSWORD_NAME)))
    passwordInput.click()
    passwordInput.send_keys(password)
    loginButton = self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_XPATH)
    loginButton.click()
    errorMessage1 = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.EMPTYMESSAGE_XPATH_1))) 
    assert errorMessage1.text == c.EMPTYMESSAGE
    errorMessage2 = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.EMPTYMESSAGE_XPATH_2))) 
    assert errorMessage2.text == c.EMPTYMESSAGE
    self.driver.close()

  @pytest.mark.parametrize("email,password",getData())
  def test_valid_Login(self,email,password):
    emailInput = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.EMAIL_NAME)))
    emailInput.click()
    emailInput.send_keys(email)
    passwordInput = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.PASSWORD_NAME)))
    passwordInput.click()
    passwordInput.send_keys(password)
    loginButton = self.driver.find_element(By.XPATH, c.LOGIN_BUTTON_XPATH)
    loginButton.click()
    validLogin = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, c.VALID_LOGIN_XPATH)))
    assert validLogin.text == c.VALID_LOGIN
    self.driver.save_screenshot("./valid.png")
    self.driver.close()
  

  @pytest.mark.parametrize("email,password",getData1()) 
  def test_invalid_login(self,email,password):
   emailInput = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.EMAIL_NAME)))
   emailInput.click()
   emailInput.send_keys(email)
   passwordInput = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.PASSWORD_NAME)))
   passwordInput.click()
   passwordInput.send_keys(password)
   loginButton = self.driver.find_element(By.XPATH, c.LOGIN_BUTTON_XPATH)
   loginButton.click()
   invalidLogin = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.INVALID_LOGIN_XPATH)))
   assert invalidLogin.text == c.INVALID_LOGIN
   self.driver.save_screenshot("./invalid.png")
   self.driver.close()